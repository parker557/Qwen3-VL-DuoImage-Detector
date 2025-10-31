#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CSV -> Excel(xlsx) 转换脚本（中文编码友好）

特点：
- 自动侦测常见中文编码（utf-8 / utf-8-sig / gbk / gb18030 / big5 / cp936）
- 自动嗅探分隔符（逗号、制表、分号、竖线等），也可手动指定
- 生成 .xlsx（Unicode 原生），中文在 Excel 中可正常显示
- 可选工作表名；可开启/关闭简易的列宽自适应

用法：
python csv2xlsx.py --input /path/to/file.csv --output /path/to/file.xlsx \
    [--encoding auto] [--delimiter auto] [--sheet-name 数据] [--autowidth]

示例：
python csv2xlsx.py -i result.csv -o result.xlsx --autowidth
"""

from __future__ import annotations
import argparse
import csv
import io
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ------------------------ 工具函数 ------------------------

def detect_encoding(p: Path, user_encoding: Optional[str] = None) -> str:
    """尝试判断文件编码。优先使用用户指定；否则按常见编码依次尝试。"""
    if user_encoding and user_encoding.lower() != "auto":
        return user_encoding

    candidates = [
        "utf-8-sig",  # 带 BOM 的 UTF-8
        "utf-8",
        "gbk",
        "gb18030",
        "big5",
        "cp936",    # Windows 简体中文
    ]
    data = p.read_bytes()
    for enc in candidates:
        try:
            data.decode(enc)
            return enc
        except Exception:
            continue
    # 兜底
    return "utf-8"


def sniff_delimiter(sample_text: str, default: str = ",") -> str:
    try:
        sniffer = csv.Sniffer()
        dialect = sniffer.sniff(sample_text, delimiters=[",", "\t", ";", "|"])
        return dialect.delimiter
    except Exception:
        return default


def iter_csv_rows(path: Path, encoding: str, delimiter: Optional[str]) -> tuple[str, list[list[str]]]:
    # 读取样本用于嗅探分隔符
    with path.open("r", encoding=encoding, errors="strict") as f:
        head = f.read(4096)
    delim = delimiter if (delimiter and delimiter.lower() != "auto") else sniff_delimiter(head)

    rows: list[list[str]] = []
    with path.open("r", encoding=encoding, errors="strict", newline="") as f:
        rdr = csv.reader(f, delimiter=delim)
        for row in rdr:
            rows.append([cell for cell in row])
    return delim, rows


def write_to_xlsx(rows: list[list[str]], out_path: Path, sheet_name: str, auto_width: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    # openpyxl 限制：工作表名最长 31 字符，且不能含 \\ / * ? : [ ]
    safe_title = sheet_name[:31]
    for ch in "\\/*?:[]":
        safe_title = safe_title.replace(ch, " ")
    ws.title = safe_title

    # 写入单元格
    max_col_width = {}
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            v = val if val is not None else ""
            ws.cell(row=r_idx, column=c_idx, value=v)
            if auto_width:
                ln = len(str(v))
                max_col_width[c_idx] = max(max_col_width.get(c_idx, 0), ln)

    if auto_width and max_col_width:
        for c_idx, width in max_col_width.items():
            # 适当加一点余量，中文字符在 Excel 中宽度略大
            ws.column_dimensions[get_column_letter(c_idx)].width = min(80, max(10, width * 1.2))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))


# ------------------------ 主流程 ------------------------

def main():
    ap = argparse.ArgumentParser(description="CSV 转 Excel (xlsx) —— 中文编码友好")
    ap.add_argument("-i", "--input", required=True, help="输入 CSV 文件路径")
    ap.add_argument("-o", "--output", required=True, help="输出 XLSX 文件路径")
    ap.add_argument("-e", "--encoding", default="auto", help="CSV 编码，默认 auto 自动识别")
    ap.add_argument("-d", "--delimiter", default="auto", help="分隔符，默认 auto 自动嗅探，可指定如 ',' 或 '\t'")
    ap.add_argument("-s", "--sheet-name", default="数据", help="工作表名称，默认 '数据'")
    ap.add_argument("--autowidth", action="store_true", help="根据内容粗略自适应列宽")
    args = ap.parse_args()

    in_path = Path(args.input)
    out_path = Path(args.output)

    if not in_path.exists():
        raise SystemExit(f"输入文件不存在：{in_path}")

    enc = detect_encoding(in_path, args.encoding)
    delim, rows = iter_csv_rows(in_path, enc, args.delimiter)

    print(f"[INFO] 编码: {enc} | 分隔符: {repr(delim)} | 行数: {len(rows)}")
    write_to_xlsx(rows, out_path, args.sheet_name, auto_width=args.autowidth)
    print(f"[OK] 已生成: {out_path}")


if __name__ == "__main__":
    main()
